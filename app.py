import streamlit as st
import pandas as pd
import io
import gc
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# --- 1. CONFIGURACIÃ“N ---
st.set_page_config(page_title="Separador de Actividades", page_icon="ðŸ—ï¸", layout="wide")

st.title("ðŸ—ï¸ Separador de Actividades (Orden Original)")

# --- INICIO DEL BLOQUE DE INSTRUCCIONES ---
with st.expander("ðŸ“˜ GUÃA DE USO (Haz clic aquÃ­ para leer)", expanded=False):
    st.markdown("""
    ### ðŸ“ Pasos para procesar tu archivo:
     
    1.  **Prepara tu Excel:** AsegÃºrate de que tenga los encabezados **CLAVE** y **ASIST**.
    2.  **Sube el archivo:** Arrastra tu documento abajo.
    3.  **Procesar:** Haz clic en **ðŸš€ PROCESAR DATOS**.
    4.  **RevisiÃ³n:**
        * Si todo sale bien, verÃ¡s globos ðŸŽˆ.
        * **ALERTA ROJA ðŸš¨:** * Lun-Vie: Si suma **mÃ¡s de 12 horas**.
            * SÃ¡b-Dom: Si suma **mÃ¡s de 6 horas**.
    5.  **Descargar:** ObtÃ©n tu reporte limpio y ordenado.

    ---
    ### âš™ï¸ Funciones Clave:
    * **Orden Intacto:** Respeta el orden original de tu Excel.
    * **Comida por Turno:** Respeta 1 comida para el 1er Turno y 1 para el 2do.
    * **Detector de Fatiga DinÃ¡mico:** Ajusta el lÃ­mite en fines de semana.
    """)
# --- FIN DEL BLOQUE DE INSTRUCCIONES ---

# --- 2. LÃ“GICA MAESTRA ---
def procesar_excel_master(file_content):
    try:
        excel_file = io.BytesIO(file_content)
        
        # A) LECTURA DE ESTRUCTURA
        df_raw = pd.read_excel(excel_file, header=None, nrows=50)
        
        fila_encabezado = -1
        for i, fila in df_raw.iterrows():
            texto = fila.astype(str).str.upper()
            if texto.str.contains("CLAVE").any() and texto.str.contains("ASIST").any():
                fila_encabezado = i
                break
                
        if fila_encabezado == -1:
            return None, None, "âŒ No encontrÃ© 'CLAVE' y 'ASIST' al inicio."

        header_top = df_raw.iloc[fila_encabezado].ffill().astype(str).str.strip()
        header_bottom = df_raw.iloc[fila_encabezado + 1].fillna("").astype(str).str.strip()
        
        nombres_columnas = []
        indices_rmmal = []
        columna_comida = None
        
        for k in range(len(header_top)):
            top = header_top.iloc[k]
            bottom = header_bottom.iloc[k]
            if top == "nan": top = f"Col_{k}"
            if bottom == "nan" or bottom == "x": bottom = ""
            
            nombre_unico = f"{top}|{bottom}"
            nombres_columnas.append(nombre_unico)
            
            if "RMMAL" in top.upper(): indices_rmmal.append(nombre_unico)
            if "COMIDA" in top.upper() and "TOTAL" not in top.upper(): columna_comida = nombre_unico

        del df_raw
        gc.collect()

        # B) CARGA DE DATOS
        excel_file.seek(0)
        df = pd.read_excel(excel_file, header=None, skiprows=fila_encabezado + 2)
        df = df.iloc[:, :len(nombres_columnas)]
        df.columns = nombres_columnas
        
        # Convertir a nÃºmeros
        for col in indices_rmmal: 
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype('float32')
        if columna_comida: 
            df[columna_comida] = pd.to_numeric(df[columna_comida], errors='coerce').fillna(0).astype('float32')

        # C) DESGLOSE
        nuevas_filas = []
        records = df.to_dict('records')
        
        try:
            col_act_key = [c for c in df.columns if str(c).startswith("Act|")][0]
            col_turno_key = [c for c in df.columns if str(c).startswith("Turno|")][0]
        except:
            return None, None, "âŒ Faltan columnas Act o Turno vacÃ­as."

        for row in records:
            cols_activas = [c for c in indices_rmmal if row[c] > 0]
            if not cols_activas: continue
            
            columna_ganadora = max(cols_activas, key=lambda k: row[k])
            
            for col_actual in cols_activas:
                fila_nueva = row.copy()
                partes = col_actual.split('|')
                
                fila_nueva[col_act_key] = partes[0]
                fila_nueva[col_turno_key] = partes[1]
                
                for c in indices_rmmal:
                    if c != col_actual: fila_nueva[c] = 0
                
                if columna_comida and col_actual != columna_ganadora:
                    fila_nueva[columna_comida] = 0
                
                nuevas_filas.append(fila_nueva)
        
        del df, records
        gc.collect()
        
        df_final = pd.DataFrame(nuevas_filas, columns=nombres_columnas)

        # --- TRUCO DE INGENIERÃA: MEMORIA FOTOGRÃFICA ðŸ“¸ ---
        df_final['__orden_original__'] = range(len(df_final))

        # D) CANDADO FINAL (POR TURNO) ðŸ”“ðŸ”’
        if columna_comida:
            c_nombre = next((c for c in df_final.columns if "NOMBRE" in c.upper()), None)
            c_fecha = next((c for c in df_final.columns if "FECHA" in c.upper()), None)
            
            if c_nombre and c_fecha:
                # 1. Calculamos horas para desempatar
                df_final['__temp_horas__'] = df_final[indices_rmmal].sum(axis=1)
                
                # 2. ORDENAMOS TEMPORALMENTE
                df_final = df_final.sort_values(
                    by=[c_nombre, c_fecha, col_turno_key, '__temp_horas__'], 
                    ascending=[True, True, True, False]
                )
                
                # 3. Borrado de comidas duplicadas
                mask_dup = df_final.duplicated(subset=[c_nombre, c_fecha, col_turno_key], keep='first')
                df_final.loc[mask_dup, columna_comida] = 0
                
                # 4. RESTAURAMOS EL ORDEN ORIGINAL ðŸ”„
                df_final = df_final.sort_values(by='__orden_original__', ascending=True)
                
                df_final.drop(columns=['__temp_horas__', '__orden_original__'], inplace=True)
            else:
                df_final.drop(columns=['__orden_original__'], inplace=True)
        else:
             df_final.drop(columns=['__orden_original__'], inplace=True)

        # E) LIMPIEZA DE FECHAS
        cols_fecha = [c for c in df_final.columns if "FECHA" in str(c).upper()]
        for col in cols_fecha:
            df_final[col] = pd.to_datetime(df_final[col], errors='coerce').dt.date

        # G) REPORTE DE EXCESO (DinÃ¡mico segÃºn dÃ­a de la semana) ðŸ‘®â€â™‚ï¸
        df_excedidos = pd.DataFrame()
        if c_nombre and c_fecha:
            # Calculamos horas totales por fila para la alerta
            df_final['__total_fila__'] = df_final[indices_rmmal].sum(axis=1) + df_final[columna_comida]
            
            # Agrupamos solo para checar
            reporte = df_final.groupby([c_nombre, c_fecha, col_turno_key])['__total_fila__'].sum().reset_index()
            
            # --- MODIFICACIÃ“N LÃ“GICA DE DÃAS (SÃ¡bados y Domingos) ---
            # Convertimos temporalmente a datetime para sacar el dÃ­a de la semana
            reporte['__temp_date__'] = pd.to_datetime(reporte[c_fecha])
            
            # 0=Lunes ... 5=SÃ¡bado, 6=Domingo
            # Si es mayor o igual a 5 (SÃ¡b o Dom), lÃ­mite es 6.0, si no, lÃ­mite es 12.0
            reporte['__limite_horas__'] = reporte['__temp_date__'].dt.dayofweek.apply(lambda x: 6.0 if x >= 5 else 12.0)
            
            # Filtramos excesos comparando contra el lÃ­mite dinÃ¡mico
            df_excedidos = reporte[reporte['__total_fila__'] > reporte['__limite_horas__']].copy()
            
            # Limpieza para mostrar al usuario
            df_excedidos = df_excedidos[[c_nombre, c_fecha, col_turno_key, '__total_fila__']]
            df_excedidos.columns = ['Nombre', 'Fecha', 'Turno', 'Horas Totales']
            # -----------------------------------
            
            df_final.drop(columns=['__total_fila__'], inplace=True)

        # F) EXPORTACIÃ“N
        output = io.BytesIO()
        df_headers = pd.DataFrame([header_top.values, header_bottom.values], columns=nombres_columnas)
        df_export = pd.concat([df_headers, df_final], axis=0)
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, header=False, sheet_name='Reporte')
        
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        
        # ESTILOS
        fill_gris = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        font_blanca = Font(color="FFFFFF", bold=True)
        fill_amarillo = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        borde = Side(style="thin", color="000000")
        caja = Border(left=borde, right=borde, top=borde, bottom=borde)
        
        letra_col_act = None
        for col_idx, cell in enumerate(ws[1], 1):
            if "Act" in str(cell.value): letra_col_act = get_column_letter(col_idx)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = caja
                if cell.row <= 2:
                    cell.fill = fill_gris; cell.font = font_blanca
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif letra_col_act and cell.column_letter == letra_col_act:
                    cell.fill = fill_amarillo
                    cell.alignment = Alignment(horizontal="left")
                
                if cell.value and str(cell.value).startswith("202") and "-" in str(cell.value):
                     cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            try:
                max_len = max(len(str(cell.value) or "") for cell in col[:50])
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)
            except: pass
            
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        
        return final_output, df_excedidos, None
        
    except Exception as e:
        return None, None, f"Error TÃ©cnico: {str(e)}"

# --- 3. INTERFAZ ---
archivo = st.file_uploader("ðŸ“‚ Cargar Excel", type=["xlsx"])

if archivo:
    if st.button("ðŸš€ PROCESAR DATOS"):
        with st.spinner("Procesando y manteniendo el orden original..."):
            bytes_data = archivo.getvalue()
            excel_resultado, df_alertas, error_msg = procesar_excel_master(bytes_data)
            
            if error_msg:
                st.error(error_msg)
            else:
                st.success("âœ… Â¡Listo! El archivo respeta el orden original.")
                
                if df_alertas is not None and not df_alertas.empty:
                    st.error(f"âš ï¸ SE DETECTARON {len(df_alertas)} CASOS DE EXCESO DE HORAS:")
                    st.write("Nota: LÃ­mite normal 12h | LÃ­mite Fin de Semana (SÃ¡b/Dom) 6h")
                    st.dataframe(df_alertas, use_container_width=True)
                else:
                    st.balloons()
                    st.info("âœ… NingÃºn turno excediÃ³ el lÃ­mite (12h Lun-Vie / 6h Fines de Semana).")

                st.download_button(
                    label="ðŸ“¥ Descargar Reporte Final",
                    data=excel_resultado,
                    file_name="Reporte_Ordenado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
